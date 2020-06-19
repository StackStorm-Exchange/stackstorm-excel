# Licensed to the StackStorm, Inc ('StackStorm') under one or more
# contributor license agreements.  See the NOTICE file distributed with
# this work for additional information regarding copyright ownership.
# The ASF licenses this file to You under the Apache License, Version 2.0
# (the "License"); you may not use this file except in compliance with
# the License.  You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and

import mock
import yaml

from excel_base_action_test_case import ExcelBaseActionTestCase

from openpyxl import load_workbook

from datetime import datetime

from delete_row import DeleteExcelRowAction


class DeleteRowsTestCase(ExcelBaseActionTestCase):
    __test__ = True
    action_cls = DeleteExcelRowAction

    SHEET_1 = [ [ "Col1", "Col2" ], [ "ro1_1", "ro1_2" ], [ "ro2_1", "ro2_2" ] ]
    SHEET_2 = [ [ "Col1", "Col2" ] ]
    _MOCK_SHEETS = {"sheet1": SHEET_1,
                    "sheet2": SHEET_2}

    def setUp(self):
        super(DeleteRowsTestCase, self).setUp()
        self._full_config = self.load_yaml('full.yaml')

    def load_yaml(self, filename):
        return yaml.safe_load(self.get_fixture_content(filename))

    @property
    def full_config(self):
        return self._full_config


    @mock.patch('openpyxl.load_workbook',
                mock.MagicMock(return_value=ExcelBaseActionTestCase.MockWorkbook(_MOCK_SHEETS, None)))
    def test_delete_row_exists(self):

        action = self.get_action_instance(self.full_config)
        result = action.run('sheet1', 'key1', True, "mock_excel.xlsx")

        self.assertIsNotNone(result)
        self.assertTrue(isinstance(result, list))
        self.assertTrue(isinstance(result[0], dict))
        self.assertEqual(result[0]['name'], 'foo')
