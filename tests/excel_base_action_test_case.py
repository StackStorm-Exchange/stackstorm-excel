# Licensed to the StackStorm, Inc ('StackStorm') under one or more
# contributor license agreements.  See the NOTICE file distributed with
# this work for additional information regarding copyright ownership.
# The ASF licenses this file to You under the Apache License, Version 2.0
# (the "License"); you may not use this file except in compliance with
# the License.  You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and

import mock

from st2tests.base import BaseActionTestCase
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell


class ExcelBaseActionTestCase(BaseActionTestCase):
    """Base class for testing the Excel pack. Includes MockWorkbook that
    mimics an Excel workbook as an array
       of cells. Only those methods needed for the tests are mocked - it is
       not a full mock of the openpyxl
       classes.
    """
    __test__ = False

    class MockCell(Cell):
        """MockCell class just overrides repr"""

        def __init__(self, worksheet, row=None, column=None, value=None,
                     style_array=None):
            super(ExcelBaseActionTestCase.MockCell, self).__init__(worksheet,
                                                                   row, column,
                                                                   value,
                                                                   style_array)

        def __repr__(self):
            return "<Cell {}>".format(self._value)

    class MockSheet(Worksheet):
        def __init__(self, sheetname, rows, **kwargs):
            ''' Mocks worksheet as an array of rows, each row is an array
                of mock cells taken from value in rows array
                :param  sheetname: Name of sheet
                :param  rows: Array of rows, each row is array of strings for
                cell values
            '''
            super(ExcelBaseActionTestCase.MockSheet, self).__init__(**kwargs)
            self.title = sheetname
            self.mockrows = []
            for i in range(len(rows)):
                newrow = []
                for j in range(len(rows[i])):
                    newrow.append(self._createCell(
                        row=i, column=j,
                        value=rows[i][j]))
                self.mockrows.append(newrow)

        def _createCell(self, row, column, value=None):
            cell = ExcelBaseActionTestCase.MockCell(self, row=row,
                                                    column=column, value=value)
            return cell

        def delete_rows(self, idx, amount=1):
            for i in range(idx - 1, idx + amount - 1):
                self.mockrows.pop(i)

        def cell(self, row, column, value=None):
            # Returns cell, or creates one if doesn't exist
            rowfound = None
            cellfound = None
            if row <= len(self.mockrows):
                rowfound = self.mockrows[row - 1]
            if not rowfound:
                for i in range(len(self.mockrows), row):
                    self.mockrows.append([])
                rowfound = self.mockrows[row - 1]
            if column <= len(rowfound):
                cellfound = rowfound[column - 1]
            if not cellfound:
                cellfound = self._createCell(
                    row=row, column=column, value=value)
                # Put empty cells before our column
                if column > (len(rowfound) + 1):
                    for i in range(column - len(rowfound) - 1):
                        rowfound.append(
                            mock.Mock(Cell(self, row, i, value=None)))
                rowfound.append(cellfound)
            return cellfound

    class MockWorkbook(object):
        def __init__(self, worksheets, new_sheet, **kwargs):
            """ Represents workbook as a dictionary of sheets -> MockWorksheet
                :param worksheets: Dictionary of sheetname to array of array
                of strings
                :param new_sheet: """
            # Convert dictionary of sheetname -> array of values to
            # dictionary of sheetname -> MockSheet
            self._mocksheets = {}
            self.encoding = "utf-8"
            for sheetname in worksheets.keys():
                self._mocksheets[
                    sheetname] = ExcelBaseActionTestCase.MockSheet(
                    sheetname, worksheets[sheetname],
                    parent=self)

        @property
        def sheetnames(self):
            return self._mocksheets.keys()

        def get_sheet_by_name(self, sheet_name):
            return self._mocksheets[sheet_name]

        def create_sheet(self, sheet_name):
            """ Creates new empty sheet for sheet_name"""
            self._mocksheets[sheet_name] = ExcelBaseActionTestCase.MockSheet(
                sheet_name, [], parent=self)
            return self._mocksheets[sheet_name]

    def setUp(self):
        super(ExcelBaseActionTestCase, self).setUp()

    @staticmethod
    def mock_file_exists(filename):
        """ Mocks os_file_exists such that returns True to file exists,
        but False for existence of lock file"""
        if "lock" in filename:
            return False
        else:
            return True

    @staticmethod
    def _util_get_column(rows, column):
        """ Returns array of values for column specified from specified array
        of array of values
            :param rows: Array of array of strings, representing the
            rows/columns of spreadsheet
            :param column: Column to return from array (starting at 0) """
        values = []
        for i in range(len(rows)):
            values.append(rows[i][column].value)
        return values
